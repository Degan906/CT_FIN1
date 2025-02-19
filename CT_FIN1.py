import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO
import datetime

# Dados de login (usuário e senha)
credenciais = {
    "henrique.degan": "12345",
    "vanessa.degan": "12345"
}

# Função para carregar os tipos da aba "Tipo"
def carregar_tipos():
    url = "https://raw.githubusercontent.com/Degan906/CT_FIN1/main/FIN_TC1.xlsx"
    try:
        df = pd.read_excel(url, sheet_name="Tipo", engine="openpyxl")
        return df["Tipo"].tolist()
    except Exception as e:
        st.error(f"Erro ao carregar os tipos: {e}")
        return []

# Função para carregar as categorias da aba "Categoria"
def carregar_categorias():
    url = "https://raw.githubusercontent.com/Degan906/CT_FIN1/main/FIN_TC1.xlsx"
    try:
        df = pd.read_excel(url, sheet_name="Categoria", engine="openpyxl")
        return df["Categorias"].tolist()
    except Exception as e:
        st.error(f"Erro ao carregar as categorias: {e}")
        return []

# Função para carregar os status da aba "Status"
def carregar_status():
    url = "https://raw.githubusercontent.com/Degan906/CT_FIN1/main/FIN_TC1.xlsx"
    try:
        df = pd.read_excel(url, sheet_name="Status", engine="openpyxl")
        return df["Status"].tolist()
    except Exception as e:
        st.error(f"Erro ao carregar os status: {e}")
        return []

# Função para registrar um novo registro na aba "Base"
def registrar_registro(tipo, categoria, data_pagamento, valor, tag, status, tipo_conta, parcelas=None):
    arquivo_excel = "FIN_TC1.xlsx"
    try:
        workbook = load_workbook(arquivo_excel)
        if "Base" not in workbook.sheetnames:
            st.error("A aba 'Base' não foi encontrada no arquivo Excel.")
            return False
        sheet = workbook["Base"]
        proxima_linha = sheet.max_row + 1
        sheet.cell(row=proxima_linha, column=1, value=tipo)
        sheet.cell(row=proxima_linha, column=2, value=categoria)
        sheet.cell(row=proxima_linha, column=3, value=data_pagamento)
        sheet.cell(row=proxima_linha, column=4, value=valor)
        sheet.cell(row=proxima_linha, column=5, value=tag)
        sheet.cell(row=proxima_linha, column=6, value=status)
        sheet.cell(row=proxima_linha, column=7, value=tipo_conta)
        sheet.cell(row=proxima_linha, column=8, value=parcelas)
        workbook.save(arquivo_excel)
        return True
    except Exception as e:
        st.error(f"Erro ao registrar o registro: {e}")
        return False

# Função para carregar os registros da aba "Base"
def carregar_registros():
    arquivo_excel = "FIN_TC1.xlsx"
    try:
        df = pd.read_excel(arquivo_excel, sheet_name="Base", engine="openpyxl")
        return df
    except Exception as e:
        st.error(f"Erro ao carregar os registros: {e}")
        return None

# Função para baixar a planilha
def baixar_planilha():
    arquivo_excel = "FIN_TC1.xlsx"
    try:
        with open(arquivo_excel, "rb") as f:
            bytes_data = f.read()
        st.download_button(
            label="Baixar Planilha",
            data=bytes_data,
            file_name="FIN_TC1.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except FileNotFoundError:
        st.error("Erro: O arquivo 'FIN_TC1.xlsx' não foi encontrado.")

# Função para verificar login
def verificar_login(usuario, senha):
    if usuario in credenciais and credenciais[usuario] == senha:
        return True
    return False

# Função para calcular a projeção financeira
def calcular_projecao(df, meses):
    hoje = datetime.date.today()
    datas_futuras = [hoje + datetime.timedelta(days=30 * i) for i in range(meses)]
    meses_formatados = [data.strftime("%b %Y") for data in datas_futuras]
    
    # Criar DataFrame vazio para a projeção
    dados_projecao = []
    
    for _, row in df.iterrows():
        tipo = row["Tipo"]
        valor = row["R$"]
        data_pagamento = row["Data de PGTO"]
        tipo_conta = row["Tipo de Conta"]
        
        if isinstance(data_pagamento, str):
            data_pagamento = datetime.datetime.strptime(data_pagamento, "%Y-%m-%d").date()
        
        linha = {"Lançamento": f"{row['Tag']} ({row['Categoria']})"}
        
        for data in datas_futuras:
            mes_ano = data.strftime("%b %Y")
            # Verifica se o lançamento é fixo ou ocorre no mês específico
            if tipo_conta.lower() == "fixa" or (data_pagamento.month == data.month and data_pagamento.year == data.year):
                linha[mes_ano] = valor if tipo.lower() == "receita" else -valor
            else:
                linha[mes_ano] = 0
        
        dados_projecao.append(linha)
    
    # Criar DataFrame final
    df_projecao = pd.DataFrame(dados_projecao)
    
    # Adicionar linha de saldo final
    saldo_final = {"Lançamento": "Saldo Final"}
    for mes in meses_formatados:
        saldo_final[mes] = df_projecao[mes].sum()
    df_projecao = pd.concat([df_projecao, pd.DataFrame([saldo_final])], ignore_index=True)
    
    return df_projecao

# Função principal do Streamlit
def main():
    st.title("Sistema de Login")
    if "logado" not in st.session_state:
        st.session_state.logado = False

    # Tela de login
    if not st.session_state.logado:
        usuario = st.text_input("Usuário")
        senha = st.text_input("Senha", type="password")
        if st.button("Entrar"):
            if verificar_login(usuario, senha):
                st.session_state.logado = True
                st.success("Login realizado com sucesso!")
                st.rerun()
            else:
                st.error("Usuário ou senha inválidos.")
    else:
        st.sidebar.title("Menu")
        opcao = st.sidebar.selectbox("Escolha uma opção", ["Início", "Criar Registro", "Listar Registros", "Baixar Planilha"])

        if opcao == "Início":
            st.header("Projeção Financeira")
            df_registros = carregar_registros()
            if df_registros is not None and not df_registros.empty:
                meses = st.slider("Selecione o número de meses para projeção", 1, 24, 6)
                df_projecao = calcular_projecao(df_registros, meses)
                st.dataframe(df_projecao)
            else:
                st.info("Nenhum registro cadastrado para gerar projeção.")

        elif opcao == "Criar Registro":
            st.header("Criar Novo Registro")
            tipos = carregar_tipos()
            categorias = carregar_categorias()
            status_list = carregar_status()
            if tipos and categorias and status_list:
                tipo = st.selectbox("Tipo", tipos)
                categoria = st.selectbox("Categoria", categorias)
                data_pagamento = st.date_input("Data de Pagamento")
                valor = st.number_input("Valor (R$)", min_value=0.0, step=0.01)
                tag = st.text_input("Tag (Label)")
                status = st.selectbox("Status", status_list)
                tipo_conta = st.selectbox("Tipo de Conta", ["Fixa", "Parcelada"])
                parcelas = None
                if tipo_conta == "Parcelada":
                    parcelas = st.number_input("Número de Parcelas", min_value=1, step=1)
                if st.button("Salvar Registro"):
                    if not tipo or not categoria or not data_pagamento or not valor or not status:
                        st.error("Todos os campos obrigatórios devem ser preenchidos.")
                    else:
                        if registrar_registro(tipo, categoria, data_pagamento, valor, tag, status, tipo_conta, parcelas):
                            st.success("Registro salvo com sucesso!")
                        else:
                            st.error("Não foi possível salvar o registro.")
            else:
                st.error("Não foi possível carregar os tipos, categorias ou status. Verifique o arquivo no repositório.")

        elif opcao == "Listar Registros":
            st.header("Registros Cadastrados")
            df_registros = carregar_registros()
            if df_registros is not None and not df_registros.empty:
                st.dataframe(df_registros)
            else:
                st.info("Nenhum registro cadastrado.")

        elif opcao == "Baixar Planilha":
            st.header("Baixar Planilha")
            st.write("Clique no botão abaixo para baixar a planilha atualizada.")
            baixar_planilha()

# Executa o aplicativo
if __name__ == "__main__":
    main()
